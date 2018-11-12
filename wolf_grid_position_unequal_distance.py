import openpyxl
import os
import numpy as np
import pygame as pg
import collections as co
import random
import pandas as pd
from pygame.locals import *
import time
import xlrd

def count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    index_list=[]
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1
            index_list.append(certain_number)

    return count,index_list

def Read_position_file(position_file_path):
    data = xlrd.open_workbook(position_file_path)
    position_data = data.sheet_by_name('Sheet')
    return position_data

def two_points_distance(point1,center):
    distance_difference = np.linalg.norm(np.array(point1) - np.array(center))
    return distance_difference

def set_random_grid(dimension):
    grid=[random.randint(0,dimension-1),
          random.randint(0, dimension - 1)]
    return grid

def grid_to_position(grid,grid_per_distance):
    position=[round((grid[0]+0.5)*grid_per_distance),round((grid[1]+0.5)*grid_per_distance)]
    return position

def position_to_grid(position,grid_per_distance):
    grid= [round(position[0]/grid_per_distance-0.5),round((position[1])/grid_per_distance-0.5)]
    return grid

def two_grid_distance(grid_1,grid_2):
    grid_distance_x = abs(grid_2[0] - grid_1[0])
    grid_distance_y = abs(grid_2[1] - grid_1[1])
    distance = grid_distance_y+grid_distance_x
    return distance

def Count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1

    return count

def dict_to_string(values={}, sep="_", prefix="", postfix=""):
    signature = prefix
    for (key, value) in values.items():
        signature=sep.join((signature, str(key), str(value)))

    signature = signature+postfix
    signature = signature.lstrip(sep)
    return signature

def set_random_point(dimension,speed_pixel_per_second):
    while True:
        point=[round((random.randint(3,dimension-4)+0.5)*speed_pixel_per_second),
               round((random.randint(3,dimension-4) + 0.5) * speed_pixel_per_second)]
        if point[0]!=0 and point[1]!=0:
            break

    return point

def detect_within_screen(point,screen_area):
    if point != None:
        flag1=point[0]>=0 and point[0]<=screen_area[0]
        flag2=point[1]>=0 and point[1]<=screen_area[1]
        within_area_flag=flag1 and flag2
    else:
        within_area_flag = False
    return within_area_flag


class Experiment(object):
    """Docstring for Experiment. """

    def __init__(self,design_values, name="expt", sub_num=-999, sub_init="test", **constants):
        """TODO: to be defined1.

        @param name TODO
        @param sub_num TODO
        @param sub_init TODO
        @param **constants TODO

        """
        self._name = name
        self._sub_num = sub_num
        self._sub_init = sub_init
        self._constants = constants
        self._values = co.OrderedDict()
        self._values["J"]=name
        self._design_values=design_values
        self._values.update(self._constants)

    def get_expt_signature(self, sep="_", prefix="expt", postfix=""):
        return dict_to_string(self._values, sep, prefix, postfix)

    def __call__(self,trial_numbers,trial=None,  writer=None):
        for index in self._design_values:
            # here is critial: each d must be passed to the trial callable
            result = trial(index)
            to_writer_dic = self._values.copy()
            to_writer_dic.update(result)
            print(to_writer_dic)
            writer(to_writer_dic)




class Trial():
    def __init__(self, pacman_grid,bean_grid,star_grid,\
                 action_space,event_space,dimension,trial_number,**constants):

        self.screen_area_grid=[dimension-1,dimension-1]
        self.trial_number=trial_number
        self.action_space = action_space
        self.event_space=event_space
        self.pacman_grid= pacman_grid
        self.bean_grid = bean_grid
        self.star_grid=star_grid
        self.constants = constants
        self.trial_index=0
        self.result = {}

    def generate_unequal_distance_pacman(self,pacman_grid,star_grid):
        relative_position=[pacman_grid[0]==star_grid[0],pacman_grid[1]==star_grid[1]]
        while True:
            if relative_position==[1,0] :
                new_pacman_grid=[pacman_grid[0]+random.choice([-1,1]),pacman_grid[1]]
            elif relative_position==[0,1]:
                new_pacman_grid=[pacman_grid[0],pacman_grid[1]+random.choice([-1,1])]
            elif relative_position==[1,1]:
                print(pacman_grid)
                print("mid_point==initial_point")
                new_pacman_grid=pacman_grid
            else:
                print("wrong initial_point")
                new_pacman_grid=pacman_grid

            if detect_within_screen(new_pacman_grid, self.screen_area_grid):
                break
        return new_pacman_grid








    def __call__(self, design_values):
        trial_index=self.trial_index
        pacman_grid= self.pacman_grid[trial_index]
        star_grid=self.star_grid[trial_index]
        if design_values==0:
            new_pacman_grid=self.generate_unequal_distance_pacman(pacman_grid,star_grid)
        else:
            new_pacman_grid=pacman_grid
        self.trial_index=self.trial_index+1
        self.result["A"] = new_pacman_grid[0]
        self.result["B"] = new_pacman_grid[1]
        self.result["C"] = self.bean_grid[trial_index][0][0]
        self.result["D"] = self.bean_grid[trial_index][0][1]
        self.result["E"] = self.bean_grid[trial_index][1][0]
        self.result["F"] = self.bean_grid[trial_index][1][1]
        self.result["G"] = self.star_grid[trial_index][0]
        self.result["H"] = self.star_grid[trial_index][1]
        if design_values==1:
            self.result["I"]="equal_distance"
        else:
            self.result["I"]="unequal_distance"
        return self.result


class Writer(object):

    """Docstring for Writer. """

    def __init__(self, path, header,replace=False):
        """TODO: to be defined1.

        @param path TODO
        @param mode TODO

        """
        self._path = path
        self.header=header
        self.index = 0
        self.data_workbook=openpyxl.Workbook()
        self.data_worksheet = self.data_workbook.active



    def __call__(self, kwargs):
        table_title=self.header
        if self.index==0:
            for col in range(len(table_title)):
                c = col + 1
                self.data_worksheet.cell(row=1, column=c).value = table_title[col]
        self.index = self.index + 1
        self.data_worksheet.append(kwargs)
        self.data_workbook.save(filename=self._path)

def main():
    pg.init()
    project_path = os.path.abspath(".")
    file_path=dict([("fixation_path",os.path.join(project_path, "images/fixation.png")),
                     ("rest_path",os.path.join(project_path, "images/rest.png")),
                     ("introduction_path" ,os.path.join(project_path, "images/introduction.png")),
                     ("positive_feedback_path",os.path.join(project_path, "images/positive_feedback.png")),
                     ("negative_feedback_path",os.path.join(project_path, "images/negative_feedback.png")),
                     ("results_path",os.path.join(project_path,"results/")),
                     ("sheep_wolf_position_path", os.path.join(project_path, "wolf_sheep_position_index.xlsx")),
                     ("position_path", os.path.join(project_path, "position.xlsx"))])
    event_space=[pg.K_UP, pg.K_DOWN,pg.K_LEFT,pg.K_RIGHT]
    speed_pixel_per_second = 30
    dimension=21
    action_space = {pg.K_UP: (0, -speed_pixel_per_second), pg.K_DOWN: (0, speed_pixel_per_second ),
                    pg.K_LEFT: (-speed_pixel_per_second , 0), pg.K_RIGHT: (speed_pixel_per_second , 0)
                    }
    trialnumber=60
    pg.event.set_allowed([pg.QUIT,pg.KEYDOWN])
    pg.event.set_blocked([pg.MOUSEBUTTONUP,pg.MOUSEMOTION,pg.MOUSEBUTTONDOWN])
    result_header = ["pacman_initial_grid_x", "pacman_initial_grid_y",
                    "bean_1_initial_grid_x", "bean_1_initial_grid_y",
                    "bean_2_initial_grid_x", "bean_2_initial_grid_y",
                     "star_grid_x","star_grid_y","condition"]

    position = Read_position_file(file_path["sheep_wolf_position_path"])
    trial_number = position.nrows - 1
    pacman_grid = [(position.cell(i + 1, 0).value,
                    position.cell(i + 1, 1).value) for i in range(trial_number)]
    bean_grid = [((position.cell(i + 1, 2).value,
                   position.cell(i + 1, 3).value),
                  (position.cell(i + 1, 4).value,
                   position.cell(i + 1, 5).value)
                  ) for i in range(trial_number)]
    star_grid = [(position.cell(i + 1, 6).value,
                  position.cell(i + 1, 7).value) for i in range(trial_number)]
    position1_name = input("Please enter position1 name:")
    position2_name = input("Please enter position2 name:")
    design_value_1=[0,1]*int(trial_number/2)
    random.shuffle(design_value_1)
    design_value_2=[not values for values in design_value_1]
    expt_1 = Experiment(design_value_1, position1_name)
    path_1 = os.path.join(file_path["results_path"], expt_1.get_expt_signature(sep="_", postfix=".xlsx"))
    trial_1 = Trial( pacman_grid, bean_grid, star_grid,\
                  action_space, event_space,dimension,trial_number)
    writer_1 = Writer(path_1, result_header, replace=False)
    print("loading1......")
    expt_1(trialnumber, trial_1, writer_1)
    print("loading2......")
    expt_2 = Experiment(design_value_2, position2_name)
    path_2 = os.path.join(file_path["results_path"], expt_2.get_expt_signature(sep="_", postfix=".xlsx"))
    trial_2 = Trial( pacman_grid, bean_grid, star_grid,\
                  action_space, event_space,dimension,trial_number)
    writer_2 = Writer(path_2, result_header, replace=False)
    expt_2(trialnumber, trial_2, writer_2)



if __name__=="__main__":
    main()







