import openpyxl
import os
import numpy as np
import pygame as pg
import collections as co
import random
import pandas as pd
from pygame.locals import *
import time
import xlrd


def Read_position_file(position_file_path):
    data = xlrd.open_workbook(position_file_path)
    position_data = data.sheet_by_name('Sheet')
    return position_data

def two_points_distance(point1,center):
    distance_difference = np.linalg.norm(np.array(point1) - np.array(center))
    return distance_difference

def set_random_grid(dimension):
    grid=[random.randint(0,dimension-1),
          random.randint(0, dimension - 1)]
    return grid

def grid_to_position(grid,grid_per_distance):
    position=[round((grid[0]+0.5)*grid_per_distance),round((grid[1]+0.5)*grid_per_distance)]
    return position

def position_to_grid(position,grid_per_distance):
    grid= [round(position[0]/grid_per_distance-0.5),round((position[1])/grid_per_distance-0.5)]
    return grid

def two_grid_distance(grid_1,grid_2):
    grid_distance_x = abs(grid_2[0] - grid_1[0])
    grid_distance_y = abs(grid_2[1] - grid_1[1])
    distance = grid_distance_y+grid_distance_x
    return distance

def Count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1

    return count

def dict_to_string(values={}, sep="_", prefix="", postfix=""):
    signature = prefix
    for (key, value) in values.items():
        signature=sep.join((signature, str(key), str(value)))

    signature = signature+postfix
    signature = signature.lstrip(sep)
    return signature

def set_random_point(dimension,speed_pixel_per_second):
    while True:
        point=[round((random.randint(3,dimension-4)+0.5)*speed_pixel_per_second),
               round((random.randint(3,dimension-4) + 0.5) * speed_pixel_per_second)]
        if point[0]!=0 and point[1]!=0:
            break

    return point

def detect_within_screen(point,screen_area):
    if point != None:
        flag1=point[0]<screen_area[0] and 0 < point[0]
        flag2=point[1]<screen_area[1] and 0 < point[1]
        within_area_flag=flag1 and flag2
    else:
        within_area_flag = False
    return within_area_flag


class Experiment(object):
    """Docstring for Experiment. """

    def __init__(self,design_values, name="expt", sub_num=-999, sub_init="test", **constants):
        """TODO: to be defined1.

        @param name TODO
        @param sub_num TODO
        @param sub_init TODO
        @param **constants TODO

        """
        self._name = name
        self._sub_num = sub_num
        self._sub_init = sub_init
        self._constants = constants
        self._values = co.OrderedDict()
        self._values['name'] = self._name
        self._values['subNum'] = self._sub_num
        self._values['subInit'] = self._sub_init
        self._design_values=design_values
        self._values.update(self._constants)

    def get_expt_signature(self, sep="_", prefix="expt", postfix=""):
        return dict_to_string(self._values, sep, prefix, postfix)

    def __call__(self, trial_number,trial=None,  writer=None):
        for index in range(trial_number):
            # here is critial: each d must be passed to the trial callable
            result = trial(index)
            to_writer_dic = self._values.copy()
            to_writer_dic ["condition"]=self._design_values[index]
            to_writer_dic.update(result)
            writer(to_writer_dic)





class Generate_bean_grid():
    def __init__(self,dimension):
        self.dimension=dimension

    def __call__(self,min_grid_distance):
        while True:
            bean1_grid=set_random_grid(self.dimension)
            bean2_grid=set_random_grid(self.dimension)
            bean_distance=two_grid_distance(bean1_grid,bean2_grid)
            if bean_distance>min_grid_distance and bean_distance%2==0:
                break
        return [bean1_grid,bean2_grid]

class Generate_pacman_grid():
    def __init__(self,bean_grid,dimension):
        self.bean_grid=bean_grid
        self.dimension=dimension

    def __call__(self,trial_index):
        bean_distance=two_grid_distance(self.bean_grid[trial_index][1],self.bean_grid[trial_index][0])
        delta_x=abs(self.bean_grid[trial_index][1][1]-self.bean_grid[trial_index][0][1])
        delta_y=abs(self.bean_grid[trial_index][0][0]-self.bean_grid[trial_index][1][0])
        if delta_y!=0:
            gradient=(self.bean_grid[trial_index][0][1]-self.bean_grid[trial_index][1][1])/(self.bean_grid[trial_index][0][0]-self.bean_grid[trial_index][1][0])
            if gradient <0:
                left_down_grid_index=[self.bean_grid[trial_index][0][0],self.bean_grid[trial_index][1][0]].\
                    index(max([self.bean_grid[trial_index][0][0],self.bean_grid[trial_index][1][0]]))
                right_up_grid_index=not left_down_grid_index
                if bean_distance/2<delta_x:
                    min_mid_point1=[self.bean_grid[trial_index][left_down_grid_index][0],
                                    self.bean_grid[trial_index][left_down_grid_index][1]+bean_distance/2]
                    min_mid_point2 = [self.bean_grid[trial_index][right_up_grid_index][0],
                                      self.bean_grid[trial_index][right_up_grid_index][1] - bean_distance / 2]
                    if random.random()<0.5:
                        pacman_grid=[random.randint(min_mid_point1[0],self.dimension-1),min_mid_point1[1]]
                        star_grid=min_mid_point1
                    else:
                        pacman_grid = [random.randint(0,min_mid_point2[0]), min_mid_point2[1]]
                        star_grid = min_mid_point2

                else:
                    min_mid_point1=[self.bean_grid[trial_index][left_down_grid_index][0]-(bean_distance/2-delta_x),
                                    self.bean_grid[trial_index][right_up_grid_index][1]]
                    min_mid_point2 = [self.bean_grid[trial_index][right_up_grid_index][0] + (bean_distance / 2 - delta_x),
                                      self.bean_grid[trial_index][left_down_grid_index][1]]
                    if random.random()<0.5:
                        pacman_grid=[min_mid_point1[0],random.randint(min_mid_point1[1],self.dimension-1)]
                        star_grid = min_mid_point1
                    else:
                        pacman_grid = [min_mid_point2[0],random.randint(0,min_mid_point2[1])]
                        star_grid = min_mid_point2

            else:
                left_up_grid_index = [self.bean_grid[trial_index][0][0], self.bean_grid[trial_index][1][0]]. \
                    index(min([self.bean_grid[trial_index][0][0], self.bean_grid[trial_index][1][0]]))
                right_down_grid_index = not left_up_grid_index
                if bean_distance/2<delta_x:
                    min_mid_point1 = [self.bean_grid[trial_index][left_up_grid_index][0] ,
                                      self.bean_grid[trial_index][left_up_grid_index][1]+bean_distance/2]
                    min_mid_point2 = [self.bean_grid[trial_index][right_down_grid_index][0],
                                      self.bean_grid[trial_index][right_down_grid_index][1]-bean_distance/2]
                    if random.random()<0.5:
                        pacman_grid=[random.randint(0,min_mid_point1[0]),min_mid_point1[1]]
                        star_grid = min_mid_point1
                    else:
                        pacman_grid = [random.randint(min_mid_point2[0],self.dimension-1),min_mid_point2[1]]
                        star_grid = min_mid_point2

                else:
                    min_mid_point1=[self.bean_grid[trial_index][left_up_grid_index][0]+bean_distance/2-delta_x,
                                    self.bean_grid[trial_index][right_down_grid_index][1]]
                    min_mid_point2 = [self.bean_grid[trial_index][right_down_grid_index][0]-(bean_distance/2-delta_x),
                                      self.bean_grid[trial_index][left_up_grid_index][1] ]
                    if random.random()<0.5:
                        pacman_grid=[min_mid_point1[0],random.randint(min_mid_point1[1],self.dimension-1)]
                        star_grid = min_mid_point1
                    else:
                        pacman_grid = [min_mid_point2[0], random.randint(0,min_mid_point2[1])]
                        star_grid = min_mid_point2
        else:
            left_grid_index = [self.bean_grid[trial_index][0][1], self.bean_grid[trial_index][1][1]]. \
                index(min([self.bean_grid[trial_index][0][1], self.bean_grid[trial_index][1][1]]))
            right_grid_index = not left_grid_index
            mid_point = [self.bean_grid[trial_index][left_grid_index][0],
                              self.bean_grid[trial_index][left_grid_index][1] + bean_distance / 2]
            if random.random()<0.5:
                pacman_grid = [random.randint(0,mid_point[0]), mid_point[1]]
            else:
                pacman_grid=[random.randint(mid_point[0],self.dimension-1), mid_point[1]]
            star_grid = mid_point
        return [pacman_grid,star_grid]



def main():
    pg.init()
    project_path = os.path.abspath(".")
    file_path=dict([("sheep_position_path", os.path.join(project_path, "sheep_position.xlsx")),
                    ("wolf_sheep_position_path", os.path.join(project_path, "wolf_sheep_position_index.xlsx"))])
    dimension=21
    trialnumber=60
    pg.event.set_allowed([pg.QUIT,pg.KEYDOWN])
    pg.event.set_blocked([pg.MOUSEBUTTONUP,pg.MOUSEMOTION,pg.MOUSEBUTTONDOWN])
    bean_position = Read_position_file(file_path["sheep_position_path"])
    trial_number = bean_position.nrows - 1
    bean_grid = [[[bean_position.cell(i + 1, 0).value,
                   bean_position.cell(i + 1, 1).value],
                  [bean_position.cell(i + 1, 2).value,
                   bean_position.cell(i + 1, 3).value]
                  ] for i in range(trial_number)]
    generate_pacman_grid = Generate_pacman_grid(bean_grid, dimension)
    pacman_grid = [generate_pacman_grid(i) for i in range(trialnumber)]
    position_header=("pacman_grid_y","pacman_grid_x",
                     "sheep1_grid_y","sheep1_grid_x",
                     "sheep2_grid_y","sheep2_grid_x",
                     "star_grid_y","star_grid_x")
    data_workbook = openpyxl.Workbook()
    data_worksheet = data_workbook.active
    table_title = position_header
    position={}
    for index in range(len(bean_grid)):
        if index == 0:
            for col in range(len(table_title)):
                c = col + 1
                data_worksheet.cell(row=1, column=c).value = table_title[col]
        position["A"]=pacman_grid[index][0][0]
        position["B"] = pacman_grid[index][0][1]
        position["C"]=bean_grid[index][0][0]
        position["D"] = bean_grid[index][0][1]
        position["E"] = bean_grid[index][1][0]
        position["F"] = bean_grid[index][1][1]
        position["G"] = pacman_grid[index][1][0]
        position["H"] = pacman_grid[index][1][1]
        data_worksheet.append(position)
        data_workbook.save(filename=file_path["wolf_sheep_position_path"])



if __name__=="__main__":
    main()





