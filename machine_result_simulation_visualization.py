import pickle
import os
import sys
import xlrd
import pygame as pg
import random
import openpyxl
import numpy as np
import collections as co
from pygame.locals import *
import time
import xlrd

def detect_within_screen(point,screen_area):
    if point != None:
        flag1=point[0]>=0 and point[0]<=screen_area[0]
        flag2=point[1]>=0 and point[1]<=screen_area[1]
        within_area_flag=flag1 and flag2
    else:
        within_area_flag = False
    return within_area_flag

def Read_position_file(position_file_path):
    data = xlrd.open_workbook(position_file_path)
    position_data = data.sheet_by_name('Sheet')
    return position_data

def grid_to_position(grid,grid_per_distance):
    position=[round((grid[0]+0.5)*grid_per_distance),round((grid[1]+0.5)*grid_per_distance)]
    return position

def position_to_grid(position,grid_per_distance):
    grid= [round(position[0]/grid_per_distance-0.5),round((position[1])/grid_per_distance-0.5)]
    return grid

def two_grid_distance(grid_1,grid_2):
    grid_distance_x = abs(grid_2[0] - grid_1[0])
    grid_distance_y = abs(grid_2[1] - grid_1[1])
    distance = grid_distance_y+grid_distance_x
    return distance

def count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    index_list=[]
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1
            index_list.append(certain_number)

    return count,index_list

def dict_to_string(values={}, sep="_", prefix="", postfix=""):
    signature = prefix
    for (key, value) in values.items():
        signature=sep.join((signature, str(key), str(value)))

    signature = signature+postfix
    signature = signature.lstrip(sep)
    return signature


class Writer(object):

    """Docstring for Writer. """

    def __init__(self, path, header,replace=False):
        """TODO: to be defined1.

        @param path TODO
        @param mode TODO

        """
        self._path = path
        self.header=header
        self.index = 0
        self.data_workbook=openpyxl.Workbook()
        self.data_worksheet = self.data_workbook.active



    def __call__(self, kwargs):
        table_title=self.header
        if self.index==0:
            for col in range(len(table_title)):
                c = col + 1
                self.data_worksheet.cell(row=1, column=c).value = table_title[col]
        self.index = self.index + 1
        self.data_worksheet.append(kwargs)
        self.data_workbook.save(filename=self._path)


class Experiment(object):
    """Docstring for Experiment. """

    def __init__(self,design_values, name="expt", sub_num=-999, sub_init="test", **constants):
        """TODO: to be defined1.

        @param name TODO
        @param sub_num TODO
        @param sub_init TODO
        @param **constants TODO

        """
        self._name = name
        self._sub_num = sub_num
        self._sub_init = sub_init
        self._constants = constants
        self._values = co.OrderedDict()
        self._values['A'] = self._name
        self._values['B'] = self._sub_num
        self._values['C'] = self._sub_init
        self._design_values=design_values
        self._values.update(self._constants)

    def get_expt_signature(self, sep="_", prefix="expt", postfix=""):
        return dict_to_string(self._values, sep, prefix, postfix)

    def __call__(self, trial_number,trial=None,  writer=None):
        for index in range(trial_number):
            # here is critial: each d must be passed to the trial callable
            result = trial(index)
            to_writer_dic = self._values.copy()
            to_writer_dic ["D"]=self._design_values[index]
            to_writer_dic.update(result)
            writer(to_writer_dic)




class Trial():
    def __init__(self, screen_trait,background_trait,bean_trait,pacman_trait,file_path,time,\
                 center_position,pacman_grid,bean_grid,\
                 action_space,speed_pixel_per_second,policy, \
                 foolish_wolf_standard,noise_probability,font_information,dimension,**constants):
        self.screen_area=[screen_trait["width"],screen_trait["height"]]
        self.screen_area_grid=[dimension-1,dimension-1]
        self.background_trait=background_trait
        self.screen = pg.display.set_mode(self.screen_area)
        self.screen_rect = self.screen.get_rect()
        self.screen_color = screen_trait['color']
        self.noise_probability=noise_probability
        self.fixation_image = pg.image.load(file_path["fixation_path"])
        self.fixation=pg.transform.scale(self.fixation_image,(80,80))
        self.fixation_rect = self.fixation.get_rect()
        self.fixation_rect.center=center_position
        self.rest_image = pg.image.load(file_path["rest_path"])
        self.rest=pg.transform.scale(self.rest_image,(300,200))
        self.rest_rect = self.fixation.get_rect()
        self.rest_rect.center=center_position
        self.introduction_image = pg.image.load(file_path["introduction_path"])
        self.introduction=pg.transform.scale(self.introduction_image,self.screen_area)
        self.introduction_rect = self.introduction.get_rect()
        self.introduction_rect.center=center_position
        self.bean_color = bean_trait["color"]
        self.bean_size = bean_trait["size"]
        self.pacman_color=pacman_trait["color"]
        self.pacman_size = pacman_trait["size"]
        self.action_space = action_space
        self.pacman_grid= pacman_grid
        self.bean_grid = bean_grid
        self.constants = constants
        self.policy=policy
        self.bean_initial_disappear_time=time["bean_initial_disappear_time_s"]
        self.feedback_time=time["feedback_time_ms"]
        self.fixation_time = time["fixation_time_ms"]
        self.blank_screen_time=time["blank_screen_time_ms"]
        self.font_information=font_information
        self.font=pg.font.SysFont(font_information["type"],font_information["size"])
        self.text_color=font_information["color"]
        self.text_size=font_information["size"]
        self.text_type = font_information["type"]
        self.rest_time=time["rest_time_trial"]
        self.foolish_wolf_standard=foolish_wolf_standard
        self.speed_pixel_per_second=speed_pixel_per_second
        self.result = {}

    def set_preview(self):
        self.screen.fill(self.screen_color)
        start_points_horizontal=[[0,(i+1)*self.screen_area[0]/self.background_trait["dimension"]]
                                 for i in range(self.background_trait["dimension"])]
        end_points_horizontal=[[self.screen_area[0],(i+1)*self.screen_area[0]/self.background_trait["dimension"]]
                               for i in range(self.background_trait["dimension"])]
        start_points_vertical=[[(i+1)*self.screen_area[1]/self.background_trait["dimension"],0]
                               for i in range(self.background_trait["dimension"])]
        end_points_vertical = [[(i + 1) * self.screen_area[1] / self.background_trait["dimension"], self.screen_area[1]] for i in
                                 range(self.background_trait["dimension"])]
        for i in range(self.background_trait["dimension"]):
            pg.draw.line(self.screen, self.background_trait["color"], start_points_horizontal[i], end_points_horizontal[i],
                         self.background_trait["width"])
            pg.draw.line(self.screen, self.background_trait["color"], start_points_vertical[i], end_points_vertical[i],
                         self.background_trait["width"])


    def draw_fixation(self,trial_index):
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)
        while trial_index % self.rest_time ==0 and trial_index != 0:
            self.screen.blit(self.rest, self.rest_rect)
            pg.display.flip()
            pg.event.set_allowed([KEYDOWN, KEYUP])
            event=pg.event.wait()
            if event.type==pg.KEYDOWN:
                if event.key == pg.K_SPACE:
                    break
        while trial_index ==0:
            self.screen.blit(self.introduction, self.introduction_rect)
            pg.display.flip()
            pg.event.set_allowed([KEYDOWN, KEYUP])
            event=pg.event.wait()
            if event.type==pg.KEYDOWN:
                if event.key == pg.K_SPACE:
                    break
        self.set_preview()
        self.screen.blit(self.fixation, self.fixation_rect)
        pg.display.flip()
        pg.time.delay(self.fixation_time)
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)

    def check_intention(self,pacman_trajectory_list,aimed_grid,bean_1_grid,bean_2_grid):
        pacman_bean1_aimed_displacement=two_grid_distance(aimed_grid,bean_1_grid)
        pacman_bean2_aimed_displacement=two_grid_distance(aimed_grid,bean_2_grid)
        pacman_bean1_initial_displacement=two_grid_distance(bean_1_grid,pacman_trajectory_list[-1])
        pacman_bean2_initial_displacement = two_grid_distance(bean_2_grid,pacman_trajectory_list[-1])
        intention_bean1= pacman_bean1_initial_displacement-pacman_bean1_aimed_displacement
        print(pacman_bean1_aimed_displacement,pacman_bean1_initial_displacement,
              pacman_bean2_aimed_displacement,pacman_bean2_initial_displacement)
        intention_bean2= pacman_bean2_initial_displacement-pacman_bean2_aimed_displacement
        if intention_bean1>intention_bean2:
            intention=1
        elif intention_bean1<intention_bean2:
            intention=2
        else:
            intention=0
        return intention


    def update_screen(self, pacman_position, bean1_position,bean2_position):
        self.set_preview()
        pg.draw.circle(self.screen, self.pacman_color, [round(pacman_position[0]),round(pacman_position[1])], self.pacman_size)
        pg.draw.circle(self.screen, self.bean_color, bean1_position, self.bean_size)
        pg.draw.circle(self.screen, self.bean_color, bean2_position, self.bean_size)
        pg.display.flip()

    def find_next_step(self,pacman_grid,pacman_trajectory_list,
                     bean1_grid, bean2_grid, step_count,noise_point_list):
        cycle=0
        while True:
            grid_initial=pacman_grid
            new_step_count=step_count
            new_noise_point_list=noise_point_list
            this_grid_policy=self.policy[(pacman_grid,(bean1_grid,bean2_grid))]
            probability=[this_grid_policy[self.action_space[i]] for i in range (len(this_grid_policy))]
            action_space=[str(a)  for a in self.action_space]
            action = np.random.choice(action_space, 1, p=probability).tolist()
            action=eval(action[0])
            aimed_grid = tuple([x + y for (x, y) in list(zip(pacman_grid,action))])
            new_step_count=step_count
            intention=self.check_intention(pacman_trajectory_list,aimed_grid,bean1_grid,bean2_grid)
            if aimed_grid !=pacman_trajectory_list[-1]:
                new_step_count=step_count+1
            if   random.random() < self.noise_probability["intention"]:
                new_noise_point_list.append(new_step_count)
                new_pacman_grid = self.add_noise(grid_initial, action)
            else:
                new_pacman_grid = aimed_grid
            if detect_within_screen(new_pacman_grid, self.screen_area_grid):
                new_pacman_grid=new_pacman_grid
                break
            else:
                new_pacman_grid=grid_initial
                cycle=cycle+1
                if cycle>self.foolish_wolf_standard["avoid_border_maximal_step"]:
                    all_event_posible_pacman_grid = [(pacman_grid[0] + action[0], pacman_grid[1] + action[1])
                                                     for action in self.action_space]
                    within_screen_flag=[detect_within_screen(grid, self.screen_area)
                                        for grid in all_event_posible_pacman_grid]
                    count,certain_index=count_certain_number_in_list(within_screen_flag,1)
                    new_pacman_grid_index=np.random.choice(certain_index,1).tolist()
                    new_pacman_grid=all_event_posible_pacman_grid[new_pacman_grid_index[0]]
                    break
        new_pacman_position=grid_to_position(new_pacman_grid,self.speed_pixel_per_second)
        return new_pacman_position,new_pacman_grid,new_step_count,new_noise_point_list

    def add_noise(self,pacman_grid, action):
        action_space = [str(a) for a in self.action_space]
        print(action_space)
        action_space.remove(str(action))
        noise_action = np.random.choice(action_space).tolist()
        noise_action = eval(noise_action)
        print(noise_action)
        noise_grid = tuple([x + y for (x, y) in list(zip(pacman_grid, noise_action))])

        return(noise_grid)



    def check_end_condition(self,pacman_grid, bean1_grid,bean2_grid):

        end_condition2_1 = self.check_eaten(pacman_grid,bean1_grid)
        end_condition2_2 = self.check_eaten(pacman_grid,bean2_grid)
        end =  end_condition2_1 or end_condition2_2

        if end_condition2_1:
            exit_condition = "bean_1_eaten"
        elif end_condition2_2:
            exit_condition = "bean_2_eaten"
        else:
            exit_condition=False
        return end,exit_condition



    def give_feedback(self,pacman_position, bean1_position, bean2_position, exit_condition):
        self.set_preview()
        pg.draw.circle(self.screen, self.pacman_color, [round(pacman_position[0]),round(pacman_position[1])], self.pacman_size)
        if exit_condition=="bean_1_eaten":
            pg.draw.circle(self.screen, [255,215,0], bean1_position, self.bean_size)
        else:
            pg.draw.circle(self.screen, self.bean_color, bean1_position, self.bean_size)

        if exit_condition=="bean_2_eaten":
            pg.draw.circle(self.screen, [255, 215, 0], bean2_position, self.bean_size)
        else:
            pg.draw.circle(self.screen, self.bean_color, bean2_position, self.bean_size)
        pg.display.flip()
        pg.time.delay(self.feedback_time)
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)
        pg.display.flip()
        pg.time.delay(self.feedback_time)
        self.set_preview()
        pg.display.flip()

    def set_bean_disappear_probability(self, bean_disappear_flag,disappear_time,initial_time):
        bean_disappear_update_flag=bean_disappear_flag
        for i in range(len(bean_disappear_flag)):
            if bean_disappear_flag[i] == False:
                bean_disappear_update_flag[i]=random.random() < self.bean_disappear_probability
                if bean_disappear_update_flag:
                    disappear_time['bean_'+str(i+1)+'_disappear_time'] = time.time()-initial_time

        return bean_disappear_update_flag,disappear_time

    def check_eaten(self,pacman_grid,bean_grid):
        if two_grid_distance(pacman_grid,bean_grid)==0 :
            check_eaten=True
        else:
            check_eaten=False
        return check_eaten

    def update_trajectory(self, trajectory_list,pacman_position):
        if pacman_position!=trajectory_list[-1]:
            trajectory_list.append(pacman_position)

        return trajectory_list

    def __call__(self, trial_index):
        pacman_grid= self.pacman_grid[trial_index]
        bean1_grid =self.bean_grid[trial_index][0]
        bean2_grid = self.bean_grid[trial_index][1]
        pacman_position = grid_to_position(pacman_grid, self.speed_pixel_per_second)
        bean1_position = grid_to_position(bean1_grid, self.speed_pixel_per_second)
        bean2_position = grid_to_position(bean2_grid, self.speed_pixel_per_second)
        pacman_trajectory_list=[pacman_grid]
        self.draw_fixation(trial_index)
        exit_flag = False
        step_count=0
        noise_point_list=[]
        while not exit_flag:
            self.update_screen(pacman_position,bean1_position,bean2_position)
            while True:
                event = pg.event.wait()
                if event.type == pg.KEYDOWN:
                    if event.key == pg.K_SPACE:
                        break
            pacman_position,pacman_grid,step_count,noise_point_list =\
                self.find_next_step( pacman_grid,pacman_trajectory_list,
                                     bean1_grid,bean2_grid,step_count,noise_point_list)
            pacman_trajectory_list = self.update_trajectory( pacman_trajectory_list, pacman_grid)
            exit_flag,exit_condition = self.check_end_condition( pacman_grid, bean1_grid, bean2_grid)
            pg.event.set_allowed([KEYDOWN, KEYUP])

        self.give_feedback(pacman_position, bean1_position, bean2_position, exit_condition)
        self.result["E"] = self.pacman_grid[trial_index][0]
        self.result["F"] = self.pacman_grid[trial_index][1]
        self.result["G"] = self.bean_grid[trial_index][0][0]
        self.result["H"] = self.bean_grid[trial_index][0][1]
        self.result["I"] = self.bean_grid[trial_index][1][0]
        self.result["J"] = self.bean_grid[trial_index][1][1]
        self.result["K"] = exit_condition
        self.result["L"] = str(noise_point_list)
        self.result["M"] = str(pacman_trajectory_list)
        self.result["N"] = len(pacman_trajectory_list) - 1
        return self.result


def main():
    pg.init()
    project_path = os.path.abspath(".")
    file_path = dict([("fixation_path",os.path.join(project_path, "images/fixation.png")),
                     ("rest_path",os.path.join(project_path, "images/rest.png")),
                     ("introduction_path" ,os.path.join(project_path, "images/introduction.png")),
                     ("results_path",os.path.join(project_path,"results/")),
                     ("sheep_wolf_position_path", os.path.join(project_path, "images/position_index.xlsx")),
                      ("machine_policy_path",os.path.join(project_path, "machine_results/noise_0.1_sheep_states_two_policy.pkl"))])
    picklefile = open(file_path["machine_policy_path"], 'rb')
    policy = pickle.load(picklefile, encoding='iso-8859-1')
    position = Read_position_file(file_path["sheep_wolf_position_path"])
    trial_number = position.nrows - 1
    pacman_grid=[(position.cell(i + 1, 0).value,
                   position.cell(i + 1, 1).value)for i in range(trial_number)]
    bean_grid = [((position.cell(i + 1, 2).value,
                   position.cell(i + 1, 3).value),
                  (position.cell(i + 1, 4).value,
                   position.cell(i + 1, 5).value)
                  ) for i in range(trial_number)]
    event_space = [pg.K_UP, pg.K_DOWN, pg.K_LEFT, pg.K_RIGHT]
    speed_pixel_per_second = 30
    dimension = 21
    action_space =[(0,1),(0,-1),(1,0),(-1,0)]
    foolish_wolf_standard = {"maximal_step": 200, "avoid_border_maximal_step": 150}
    noise_probability = {"intention": 0.1}
    screen_trait = dict([("color", [230, 230, 230]), ("width", speed_pixel_per_second * dimension),
                         ("height", speed_pixel_per_second * dimension)])
    background_trait = dict([("color", [0, 0, 0]), ("width", 1), ("dimension", 21)])
    center_position = (screen_trait["width"] / 2, screen_trait["height"] / 2)
    trialnumber = 60
    bean_trait = dict([("color", [0, 0, 0]), ("size", 10),
                       ("bean_disappear_probability", 1)])
    pacman_trait = dict([("color", [255, 48, 48]), ("size", 10)])
    time = dict([("bean_initial_disappear_time_s", 1), ("fixation_time_ms", 1300), ("feedback_time_ms", 1000),
                 ("blank_screen_time_ms", 1000), ("rest_time_trial", trialnumber / 2)])
    font_information = dict([("color", [30, 30, 30]), ("size", 30), ("type", "arial")])
    result_header = ['name', "subNum", "subInit", "design_values"
                      "pacman_initial_grid_x", "pacman_initial_grid_y",
                     "bean_1_initial_grid_x", "bean_1_initial_grid_y",
                     "bean_2_initial_grid_x", "bean_2_initial_grid_y",
                     "end_condition_reason", "total_trial_time",
                     "noise_point_list", "pacman_trajectory", "reaction_time"]
    bean_position = Read_position_file(file_path["sheep_wolf_position_path"])
    participant_name = input("Please enter your name:")
    participant_order = input("Please enter your order:")
    test_or_experiment = input("test or experiment? ")
    design_value = list(range(trialnumber))
    random.shuffle(design_value)
    expt = Experiment(design_value, participant_name, participant_order, test_or_experiment)
    path = os.path.join(file_path["results_path"], expt.get_expt_signature(sep="_", postfix=".xlsx"))
    trial = Trial(screen_trait, background_trait, bean_trait, pacman_trait, file_path, time, \
                  center_position, pacman_grid, bean_grid, \
                  action_space,speed_pixel_per_second, policy, \
                  foolish_wolf_standard,noise_probability, font_information,dimension)
    writer = Writer(path, result_header, replace=False)
    print("loading......")
    expt(trialnumber, trial, writer)



if __name__ == "__main__":
    main()