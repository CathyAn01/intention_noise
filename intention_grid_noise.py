import openpyxl
import os
import numpy as np
import pygame as pg
import collections as co
import random
import pandas as pd
from pygame.locals import *
import time
import xlrd

def count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    index_list=[]
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1
            index_list.append(certain_number)

    return count,index_list

def Read_position_file(position_file_path):
    data = xlrd.open_workbook(position_file_path)
    position_data = data.sheet_by_name('Sheet')
    return position_data


def set_random_grid(dimension):
    grid=[random.randint(0,dimension-1),
          random.randint(0, dimension - 1)]
    return grid

def grid_to_position(grid,grid_per_distance):
    position=[round((grid[0]+0.5)*grid_per_distance),round((grid[1]+0.5)*grid_per_distance)]
    return position

def position_to_grid(position,grid_per_distance):
    grid= [round(position[0]/grid_per_distance-0.5),round((position[1])/grid_per_distance-0.5)]
    return grid

def two_grid_distance(grid_1,grid_2):
    grid_distance_x = abs(grid_2[0] - grid_1[0])
    grid_distance_y = abs(grid_2[1] - grid_1[1])
    distance = grid_distance_y+grid_distance_x
    return distance

def Count_certain_number_in_list(list_to_manipulate, certain_number):
    count = 0
    number_index_initial = 0
    for number in list_to_manipulate:
        if certain_number == number:
            count = count + 1

    return count

def dict_to_string(values={}, sep="_", prefix="", postfix=""):
    signature = prefix
    for (key, value) in values.items():
        signature=sep.join((signature, str(key), str(value)))

    signature = signature+postfix
    signature = signature.lstrip(sep)
    return signature

def set_random_point(dimension,speed_pixel_per_second):
    while True:
        point=[round((random.randint(3,dimension-4)+0.5)*speed_pixel_per_second),
               round((random.randint(3,dimension-4) + 0.5) * speed_pixel_per_second)]
        if point[0]!=0 and point[1]!=0:
            break

    return point

def detect_within_screen(point,screen_area):
    if point != None:
        flag1=point[0]>=0 and point[0]<=screen_area[0]
        flag2=point[1]>=0 and point[1]<=screen_area[1]
        within_area_flag=flag1 and flag2
    else:
        within_area_flag = False
    return within_area_flag


class Experiment(object):
    """Docstring for Experiment. """

    def __init__(self,design_values, name="expt", sub_num=-999, sub_init="test", **constants):
        """TODO: to be defined1.

        @param name TODO
        @param sub_num TODO
        @param sub_init TODO
        @param **constants TODO

        """
        self._name = name
        self._sub_num = sub_num
        self._sub_init = sub_init
        self._constants = constants
        self._values = co.OrderedDict()
        self._values['A'] = self._name
        self._values['B'] = self._sub_num
        self._values['C'] = self._sub_init
        self._design_values=design_values
        self._values.update(self._constants)

    def get_expt_signature(self, sep="_", prefix="expt", postfix=""):
        return dict_to_string(self._values, sep, prefix, postfix)

    def __call__(self, trial_number,trial=None,  writer=None):
        for index in range(trial_number):
            # here is critial: each d must be passed to the trial callable
            result = trial(index)
            to_writer_dic = self._values.copy()
            to_writer_dic ["D"]=self._design_values[index]
            to_writer_dic.update(result)
            writer(to_writer_dic)




class Trial():
    def __init__(self, screen_trait,background_trait,bean_trait,pacman_trait,file_path,time,\
                 center_position,pacman_grid,bean_grid,star_grid,\
                 action_space,event_space,speed_pixel_per_second,\
                 noise_probability,font_information,score_information,count_step,dimension,**constants):
        self.screen_area=[screen_trait["width"],screen_trait["height"]]
        self.screen_area_grid=[dimension-1,dimension-1]
        self.background_trait=background_trait
        self.screen = pg.display.set_mode(self.screen_area)
        self.screen_rect = self.screen.get_rect()
        self.screen_color = screen_trait['color']
        self.noise_probability=noise_probability
        self.fixation_image = pg.image.load(file_path["fixation_path"])
        self.fixation=pg.transform.scale(self.fixation_image,(80,80))
        self.fixation_rect = self.fixation.get_rect()
        self.fixation_rect.center=center_position
        self.rest_image = pg.image.load(file_path["rest_path"])
        self.rest=pg.transform.scale(self.rest_image,(300,200))
        self.rest_rect = self.fixation.get_rect()
        self.rest_rect.center=center_position
        self.introduction_image = pg.image.load(file_path["introduction_path"])
        self.introduction=pg.transform.scale(self.introduction_image,self.screen_area)
        self.introduction_rect = self.introduction.get_rect()
        self.introduction_rect.center=center_position
        self.bean_color = bean_trait["color"]
        self.bean_size = bean_trait["size"]
        self.pacman_color=pacman_trait["color"]
        self.pacman_size = pacman_trait["size"]
        self.action_space = action_space
        self.event_space=event_space
        self.pacman_grid= pacman_grid
        self.bean_grid = bean_grid
        self.star_grid=star_grid
        self.constants = constants
        self.bean_initial_disappear_time=time["bean_initial_disappear_time_s"]
        self.feedback_time=time["feedback_time_ms"]
        self.fixation_time = time["fixation_time_ms"]
        self.blank_screen_time=time["blank_screen_time_ms"]
        self.font_information=font_information
        self.score_information=score_information
        self.font=pg.font.SysFont(font_information["type"],font_information["size"])
        self.text_color=font_information["color"]
        self.text_size=font_information["size"]
        self.text_type = font_information["type"]
        self.initial_score=self.score_information["initial_score"]
        self.score_position=self.score_information["position"]
        self.rest_time=time["rest_time_trial"]
        self.count_step=count_step
        self.speed_pixel_per_second=speed_pixel_per_second
        self.result = {}

    def set_preview(self):
        self.screen.fill(self.screen_color)
        start_points_horizontal=[[0,(i+1)*self.screen_area[0]/self.background_trait["dimension"]]
                                 for i in range(self.background_trait["dimension"])]
        end_points_horizontal=[[self.screen_area[0],(i+1)*self.screen_area[0]/self.background_trait["dimension"]]
                               for i in range(self.background_trait["dimension"])]
        start_points_vertical=[[(i+1)*self.screen_area[1]/self.background_trait["dimension"],0]
                               for i in range(self.background_trait["dimension"])]
        end_points_vertical = [[(i + 1) * self.screen_area[1] / self.background_trait["dimension"], self.screen_area[1]] for i in
                                 range(self.background_trait["dimension"])]
        for i in range(self.background_trait["dimension"]):
            pg.draw.line(self.screen, self.background_trait["color"], start_points_horizontal[i], end_points_horizontal[i],
                         self.background_trait["width"])
            pg.draw.line(self.screen, self.background_trait["color"], start_points_vertical[i], end_points_vertical[i],
                         self.background_trait["width"])


    def draw_fixation(self,trial_index):
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)
        while trial_index % self.rest_time ==0 and trial_index != 0:
            self.screen.blit(self.rest, self.rest_rect)
            pg.display.flip()
            pg.event.set_allowed([KEYDOWN, KEYUP])
            event=pg.event.wait()
            if event.type==pg.KEYDOWN:
                if event.key == pg.K_SPACE:
                    break
        while trial_index ==0:
            self.screen.blit(self.introduction, self.introduction_rect)
            pg.display.flip()
            pg.event.set_allowed([KEYDOWN, KEYUP])
            event=pg.event.wait()
            if event.type==pg.KEYDOWN:
                if event.key == pg.K_SPACE:
                    break
        self.set_preview()
        self.screen.blit(self.fixation, self.fixation_rect)
        pg.display.flip()
        pg.time.delay(self.fixation_time)
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)

    def check_intention(self,pacman_trajectory_list,aimed_grid,bean_1_grid,bean_2_grid):
        pacman_bean1_aimed_displacement=two_grid_distance(aimed_grid,bean_1_grid)
        pacman_bean2_aimed_displacement=two_grid_distance(aimed_grid,bean_2_grid)
        pacman_bean1_initial_displacement=two_grid_distance(bean_1_grid,pacman_trajectory_list[-1])
        pacman_bean2_initial_displacement = two_grid_distance(bean_2_grid,pacman_trajectory_list[-1])
        intention_bean1= pacman_bean1_initial_displacement-pacman_bean1_aimed_displacement
        intention_bean2= pacman_bean2_initial_displacement-pacman_bean2_aimed_displacement
        if intention_bean1>intention_bean2:
            intention=1
        elif intention_bean1<intention_bean2:
            intention=2
        else:
            intention=0
        return intention


    def update_screen(self, pacman_position, bean1_position,bean2_position):
        self.set_preview()
        pg.draw.circle(self.screen, self.pacman_color, [round(pacman_position[0]),round(pacman_position[1])], self.pacman_size)
        pg.draw.circle(self.screen, self.bean_color, bean1_position, self.bean_size)
        pg.draw.circle(self.screen, self.bean_color, bean2_position, self.bean_size)
        pg.display.flip()

    def check_events(self, event, pacman_position, pacman_grid,pacman_trajectory_list,
                     bean1_grid, bean2_grid, step_count,noise_point_list,intention_list):
        grid_initial=pacman_grid
        new_step_count=step_count
        if event.type == pg.KEYDOWN and event.key in self.event_space:
            new_step_count=step_count+1
            aimed_grid = [x + y for (x, y) in list(zip(pacman_grid,
                                                     [self.action_space[event.key][0] / self.speed_pixel_per_second,
                                                      self.action_space[event.key][1] / self.speed_pixel_per_second]))]
            intention=self.check_intention(pacman_trajectory_list,aimed_grid,bean1_grid,bean2_grid)
            intention_list.append(intention)
            if  random.random() < self.noise_probability["intention"]:
                noise_point_list.append(new_step_count)
                new_pacman_grid=self.add_noise(grid_initial,event)
            else:
                new_pacman_grid = aimed_grid
        else:
            new_pacman_grid=grid_initial
        if detect_within_screen(new_pacman_grid, self.screen_area_grid):
            new_pacman_grid=new_pacman_grid
        else:
            new_pacman_grid=grid_initial
        new_pacman_position=grid_to_position(new_pacman_grid,self.speed_pixel_per_second)
        return new_pacman_position,new_pacman_grid,new_step_count,noise_point_list,intention_list

    def add_noise(self,pacman_grid, event):
        event_space=self.event_space.copy()
        event_space.remove(event.key)
        all_event_posible_pacman_grid = [[pacman_grid[0] + self.action_space[per_event][0]// self.speed_pixel_per_second,
                                          pacman_grid[1] + self.action_space[per_event][1] / self.speed_pixel_per_second
                                          ] for per_event in event_space]
        all_event_posible_pacman_grid_str = [str(a) for a in all_event_posible_pacman_grid]
        new_pacman_grid_str = np.random.choice(all_event_posible_pacman_grid_str)
        new_pacman_grid=eval(new_pacman_grid_str)
        return(new_pacman_grid)

    def update_trajectory(self,event, trajectory_list,pacman_position):
        if event.type == pg.KEYDOWN and event.key in self.event_space:
            trajectory_list.append(pacman_position)
        return trajectory_list

    def update_response_time(self,event,response_current_time_list):
        if event.type == pg.KEYDOWN and event.key in self.event_space:
            now=time.time()
            response_current_time_list.append(now)
        return response_current_time_list

    def check_end_condition(self, event, pacman_grid, bean1_grid,bean2_grid):
        if event.type == pg.QUIT:
            end_condition1 = True
        else:
            end_condition1 = False

        end_condition2_1 = self.check_eaten(pacman_grid,bean1_grid)
        end_condition2_2 = self.check_eaten(pacman_grid,bean2_grid)
        end = end_condition1 or end_condition2_1 or end_condition2_2
        if end_condition1:
            exit_condition = "QUIT"
        elif end_condition2_1:
            exit_condition = "bean_1_eaten"
        elif end_condition2_2:
            exit_condition = "bean_2_eaten"
        else:
            exit_condition=False
        return end,exit_condition



    def give_feedback(self,pacman_position, bean1_position, bean2_position, exit_condition):
        self.set_preview()
        pg.draw.circle(self.screen, self.pacman_color, [round(pacman_position[0]),round(pacman_position[1])], self.pacman_size)
        if exit_condition=="bean_1_eaten":
            pg.draw.circle(self.screen, [255,215,0], bean1_position, self.bean_size)
        else:
            pg.draw.circle(self.screen, self.bean_color, bean1_position, self.bean_size)

        if exit_condition=="bean_2_eaten":
            pg.draw.circle(self.screen, [255, 215, 0], bean2_position, self.bean_size)
        else:
            pg.draw.circle(self.screen, self.bean_color, bean2_position, self.bean_size)
        pg.display.flip()
        pg.time.delay(self.feedback_time)
        self.set_preview()
        pg.display.flip()
        pg.time.delay(self.blank_screen_time)
        pg.display.flip()
        pg.time.delay(self.feedback_time)
        self.set_preview()
        pg.display.flip()

    def set_bean_disappear_probability(self, bean_disappear_flag,disappear_time,initial_time):
        bean_disappear_update_flag=bean_disappear_flag
        for i in range(len(bean_disappear_flag)):
            if bean_disappear_flag[i] == False:
                bean_disappear_update_flag[i]=random.random() < self.bean_disappear_probability
                if bean_disappear_update_flag:
                    disappear_time['bean_'+str(i+1)+'_disappear_time'] = time.time()-initial_time

        return bean_disappear_update_flag,disappear_time

    def check_eaten(self,pacman_grid,bean_grid):
        if two_grid_distance(pacman_grid,bean_grid)==0 :
            check_eaten=True
        else:
            check_eaten=False
        return check_eaten

    def check_first_intention(self, intention_list):
        if 1 in intention_list:
            intention_index_1=intention_list.index(1)
        else:
            intention_index_1=999
        if 2 in intention_list:
            intention_index_2=intention_list.index(2)
        else:
            intention_index_2=999
        if intention_index_1<intention_index_2:
            first_intention=1
        else:
            first_intention=2
        return first_intention

    def check_last_intention(self,exit_condition):
        if exit_condition == "bean_1_eaten":
            last_intention = 1
        elif exit_condition == "bean_2_eaten":
            last_intention = 2
        else:
            last_intention = 0
        return last_intention

    def __call__(self, trial_index):
        pacman_grid= self.pacman_grid[trial_index]
        bean1_grid =self.bean_grid[trial_index][0]
        bean2_grid = self.bean_grid[trial_index][1]
        pacman_position = grid_to_position(pacman_grid, self.speed_pixel_per_second)
        bean1_position = grid_to_position(bean1_grid, self.speed_pixel_per_second)
        bean2_position = grid_to_position(bean2_grid, self.speed_pixel_per_second)
        pacman_trajectory_list=[pacman_grid]
        self.draw_fixation(trial_index)
        exit_flag = False
        initial_trial_time = time.time()
        response_current_time_list = [initial_trial_time]
        pg.event.set_allowed([KEYDOWN, KEYUP])
        step_count=0
        noise_point_list=[]
        intention_list=[]
        while not exit_flag:
            self.update_screen(pacman_position, bean1_position, bean2_position)
            for event in pg.event.get():
                pacman_position,pacman_grid,step_count,noise_point_list,intention_list\
                        = self.check_events(event, pacman_position,pacman_grid,\
                        pacman_trajectory_list,bean1_grid,bean2_grid,step_count,\
                                            noise_point_list,intention_list)
                pacman_trajectory_list=self.update_trajectory(event,pacman_trajectory_list,pacman_grid)
                response_current_time_list=self.update_response_time(event,response_current_time_list)
                exit_flag,exit_condition = self.check_end_condition(event, pacman_grid, bean1_grid, bean2_grid)

        pg.event.set_blocked([KEYDOWN, KEYUP])
        reaction_time=[response_current_time_list[i+1]-response_current_time_list[i] for i in range(len(response_current_time_list)-1)]
        self.give_feedback(pacman_position, bean1_position, bean2_position, exit_condition)
        end_trial_time=time.time()
        first_intention = self.check_first_intention(intention_list)
        last_intention = self.check_last_intention(exit_condition)
        self.result["E"] = self.pacman_grid[trial_index][0]
        self.result["F"] = self.pacman_grid[trial_index][1]
        self.result["G"] = self.bean_grid[trial_index][0][0]
        self.result["H"] = self.bean_grid[trial_index][0][1]
        self.result["I"] = self.bean_grid[trial_index][1][0]
        self.result["J"] = self.bean_grid[trial_index][1][1]
        self.result["K"] = exit_condition
        self.result["L"]=str(noise_point_list)
        self.result["M"] = str(pacman_trajectory_list)
        self.result["N"] = len(pacman_trajectory_list)-1
        self.result["O"] = first_intention
        self.result["P"] = last_intention
        self.result["Q"] = str(intention_list)
        self.result["R"] = str(reaction_time)
        self.result["S"] = end_trial_time - initial_trial_time
        self.result["T"] = self.star_grid[trial_index][0]
        self.result["U"] = self.star_grid[trial_index][1]

        return self.result


class Writer(object):

    """Docstring for Writer. """

    def __init__(self, path, header,replace=False):
        """TODO: to be defined1.

        @param path TODO
        @param mode TODO

        """
        self._path = path
        self.header=header
        self.index = 0
        self.data_workbook=openpyxl.Workbook()
        self.data_worksheet = self.data_workbook.active



    def __call__(self, kwargs):
        table_title=self.header
        if self.index==0:
            for col in range(len(table_title)):
                c = col + 1
                self.data_worksheet.cell(row=1, column=c).value = table_title[col]
        self.index = self.index + 1
        self.data_worksheet.append(kwargs)
        self.data_workbook.save(filename=self._path)

def main():
    pg.init()
    project_path = os.path.abspath(".")
    file_path=dict([("fixation_path",os.path.join(project_path, "images/fixation.png")),
                     ("rest_path",os.path.join(project_path, "images/rest.png")),
                     ("introduction_path" ,os.path.join(project_path, "images/introduction.png")),
                     ("positive_feedback_path",os.path.join(project_path, "images/positive_feedback.png")),
                     ("negative_feedback_path",os.path.join(project_path, "images/negative_feedback.png")),
                     ("results_path",os.path.join(project_path,"results/")),
                     ("sheep_wolf_position_path", os.path.join(project_path, "images/position_index.xlsx")),
                     ("position_path", os.path.join(project_path, "position.xlsx"))])
    event_space=[pg.K_UP, pg.K_DOWN,pg.K_LEFT,pg.K_RIGHT]
    speed_pixel_per_second = 30
    dimension=21
    action_space = {pg.K_UP: (0, -speed_pixel_per_second), pg.K_DOWN: (0, speed_pixel_per_second ),
                    pg.K_LEFT: (-speed_pixel_per_second , 0), pg.K_RIGHT: (speed_pixel_per_second , 0)
                    }
    noise_probability={"intention":0.1}
    screen_trait = dict([("color", [230,230,230]), ("width", speed_pixel_per_second*dimension),
                         ("height", speed_pixel_per_second*dimension)])
    background_trait=dict([("color", [0, 0, 0]), ("width", 1), ("dimension",21)])
    center_position=(screen_trait["width"]/2,screen_trait["height"]/2)
    trialnumber=60
    bean_trait = dict([("color", [0, 0, 0]), ("size",10),
                       ("bean_disappear_probability", 1)])
    pacman_trait = dict([("color", [255, 48, 48]), ("size", 10)])
    time = dict([("bean_initial_disappear_time_s", 1), ("fixation_time_ms", 1300), ("feedback_time_ms", 1000),
                 ("blank_screen_time_ms",1000),("rest_time_trial",trialnumber/2 )])
    font_information=dict([("color", [30,30,30]), ("size", 30),("type", "arial")])
    score_information=dict([("position", [4/5*screen_trait["width"],5/6*screen_trait["height"]]), ("initial_score", 0)])
    count_step=[0,1]
    pg.event.set_allowed([pg.QUIT,pg.KEYDOWN])
    pg.event.set_blocked([pg.MOUSEBUTTONUP,pg.MOUSEMOTION,pg.MOUSEBUTTONDOWN])
    result_header = ['name', "subNum", "subInit", "design_values"
                     "pacman_initial_grid_x", "pacman_initial_grid_y",
                    "bean_1_initial_grid_x", "bean_1_initial_grid_y",
                    "bean_2_initial_grid_x", "bean_2_initial_grid_y",
                    "end_condition_reason",  "noise_point_list",
                     "pacman_trajectory","step","second_order_intention",
                     "last_intention","second_order_intention_list",
                     "reaction_time","total_trial_time",
                     "star_grid_x","star_grid_y"]

    position = Read_position_file(file_path["sheep_wolf_position_path"])
    trial_number = position.nrows - 1
    pacman_grid = [(position.cell(i + 1, 0).value,
                    position.cell(i + 1, 1).value) for i in range(trial_number)]
    bean_grid = [((position.cell(i + 1, 2).value,
                   position.cell(i + 1, 3).value),
                  (position.cell(i + 1, 4).value,
                   position.cell(i + 1, 5).value)
                  ) for i in range(trial_number)]
    star_grid = [(position.cell(i + 1, 6).value,
                  position.cell(i + 1, 7).value) for i in range(trial_number)]
    participant_name = input("Please enter your name:")
    participant_order = input("Please enter your order:")
    test_or_experiment = input("test or experiment? ")
    design_value=list(range(trialnumber))
    random.shuffle(design_value)
    expt = Experiment(design_value, participant_name, participant_order, test_or_experiment)
    path = os.path.join(file_path["results_path"], expt.get_expt_signature(sep="_", postfix=".csv"))
    trial = Trial(screen_trait, background_trait, bean_trait, pacman_trait, file_path, time, \
                  center_position, pacman_grid, bean_grid, star_grid,\
                  action_space, event_space, speed_pixel_per_second, \
                  noise_probability, font_information, score_information, count_step,dimension)
    writer = Writer(path, result_header, replace=False)
    print("loading......")
    expt(trialnumber, trial, writer)



if __name__=="__main__":
    main()







